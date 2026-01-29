"""Парсер Excel файлов табелей РТБ"""
import logging
from datetime import datetime, date
from decimal import Decimal
from typing import Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class TimeSheetExcelParseError(Exception):
    """Ошибка парсинга Excel табеля"""
    pass


class TimeSheetExcelParser:
    """
    Парсер Excel файлов табелей рабочего времени
    
    Ожидаемый формат:
    - Строка 1: Заголовок "Табель рабочего времени бригады"
    - Строка 2: Бригада: [название]
    - Строка 3: Период: [дата начала] - [дата окончания]
    - Строка 4: Пустая
    - Строка 5: Заголовки столбцов
    - Строки 6+: Данные (ФИО, Дата, Объект, Часы)
    """
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.worksheet: Optional[Worksheet] = None
    
    def parse(self) -> dict:
        """
        Парсинг Excel файла
        
        Returns:
            dict: {
                'brigade_name': str,
                'period_start': date,
                'period_end': date,
                'items': [
                    {
                        'member_name': str,
                        'date': date,
                        'cost_object_name': str,
                        'hours': Decimal
                    }
                ]
            }
        """
        try:
            self.workbook = load_workbook(self.file_path, data_only=True)
            self.worksheet = self.workbook.active
            
            # Извлечение метаданных
            brigade_name = self._extract_brigade_name()
            period_start, period_end = self._extract_period()
            
            # Извлечение строк данных
            items = self._extract_items()
            
            if not items:
                raise TimeSheetExcelParseError("Не найдено ни одной строки с данными")
            
            return {
                'brigade_name': brigade_name,
                'period_start': period_start,
                'period_end': period_end,
                'items': items
            }
        
        except Exception as e:
            logger.error(f"Ошибка парсинга Excel: {e}")
            raise TimeSheetExcelParseError(f"Не удалось распарсить файл: {str(e)}")
        
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _extract_brigade_name(self) -> str:
        """Извлечение названия бригады из строки 2"""
        cell_value = self.worksheet.cell(row=2, column=1).value
        
        if not cell_value:
            raise TimeSheetExcelParseError("Не найдено название бригады (строка 2)")
        
        # Формат: "Бригада: Название"
        if ':' in str(cell_value):
            brigade_name = str(cell_value).split(':', 1)[1].strip()
        else:
            brigade_name = str(cell_value).strip()
        
        if not brigade_name:
            raise TimeSheetExcelParseError("Пустое название бригады")
        
        return brigade_name
    
    def _extract_period(self) -> tuple[date, date]:
        """Извлечение периода из строки 3"""
        cell_value = self.worksheet.cell(row=3, column=1).value
        
        if not cell_value:
            raise TimeSheetExcelParseError("Не найден период (строка 3)")
        
        # Формат: "Период: ДД.ММ.ГГГГ - ДД.ММ.ГГГГ"
        try:
            period_str = str(cell_value)
            if ':' in period_str:
                period_str = period_str.split(':', 1)[1].strip()
            
            # Разделить по дефису
            parts = [p.strip() for p in period_str.split('-')]
            if len(parts) != 2:
                raise ValueError("Неверный формат периода")
            
            # Парсинг дат
            period_start = datetime.strptime(parts[0], "%d.%m.%Y").date()
            period_end = datetime.strptime(parts[1], "%d.%m.%Y").date()
            
            if period_start > period_end:
                raise ValueError("Дата начала больше даты окончания")
            
            return period_start, period_end
        
        except Exception as e:
            raise TimeSheetExcelParseError(f"Ошибка парсинга периода: {e}")
    
    def _extract_items(self) -> list[dict]:
        """Извлечение строк данных (начиная с 6-й строки)"""
        items = []
        
        # Начинаем со строки 6 (после заголовков)
        row_num = 6
        max_empty_rows = 5  # Остановиться после 5 пустых строк подряд
        empty_count = 0
        
        while empty_count < max_empty_rows:
            # Чтение ячеек
            member_name = self.worksheet.cell(row=row_num, column=1).value
            date_value = self.worksheet.cell(row=row_num, column=2).value
            object_name = self.worksheet.cell(row=row_num, column=3).value
            hours_value = self.worksheet.cell(row=row_num, column=4).value
            
            # Проверка на пустую строку
            if not any([member_name, date_value, object_name, hours_value]):
                empty_count += 1
                row_num += 1
                continue
            
            # Сброс счётчика пустых строк
            empty_count = 0
            
            # Валидация обязательных полей
            if not member_name:
                logger.warning(f"Пропущена строка {row_num}: нет ФИО")
                row_num += 1
                continue
            
            if not date_value:
                logger.warning(f"Пропущена строка {row_num}: нет даты")
                row_num += 1
                continue
            
            if not hours_value:
                logger.warning(f"Пропущена строка {row_num}: нет часов")
                row_num += 1
                continue
            
            try:
                # Парсинг даты
                if isinstance(date_value, datetime):
                    work_date = date_value.date()
                elif isinstance(date_value, date):
                    work_date = date_value
                else:
                    work_date = datetime.strptime(str(date_value), "%d.%m.%Y").date()
                
                # Парсинг часов
                hours = Decimal(str(hours_value))
                
                if hours <= 0 or hours > 24:
                    raise ValueError(f"Некорректное количество часов: {hours}")
                
                # Добавление записи
                items.append({
                    'member_name': str(member_name).strip(),
                    'date': work_date,
                    'cost_object_name': str(object_name).strip() if object_name else "Не указан",
                    'hours': hours
                })
            
            except Exception as e:
                logger.warning(f"Ошибка парсинга строки {row_num}: {e}")
            
            row_num += 1
        
        return items
