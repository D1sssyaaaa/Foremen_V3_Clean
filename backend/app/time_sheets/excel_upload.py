"""Endpoint для загрузки табелей из Excel"""
import tempfile
import os
from datetime import datetime, timedelta
from typing import Dict, Any
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

from app.core.database import get_db
from app.auth.dependencies import require_roles, get_current_user
from app.core.models_base import UserRole
from app.models import User, CostObject, Brigade
from app.time_sheets.service import TimeSheetService
from app.time_sheets.excel_parser import TimeSheetExcelParser, TimeSheetExcelParseError
from app.time_sheets.excel_preview_service import ExcelPreviewService
from app.time_sheets.schemas import TimeSheetCreate, TimeSheetResponse, TimeSheetItemResponse

router = APIRouter()


@router.get("/template", response_class=StreamingResponse)
async def download_template(
    brigade_id: int = Query(..., description="ID бригады"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.FOREMAN, UserRole.HR_MANAGER]))
):
    """
    Скачивание шаблона табеля рабочего времени в Excel
    
    - Доступно: FOREMAN, HR_MANAGER
    - Формат: .xlsx
    - Включает пример заполнения и форматирование
    
    Структура шаблона:
    - Строка 1: Заголовок
    - Строка 2: Бригада [название]
    - Строка 3: Период [дата начала] - [дата окончания]
    - Строки 6+: Таблица с примерами (ФИО | Дата | Объект | Часы)
    """
    try:
        # Получение данных бригады
        brigade_result = await db.execute(
            select(Brigade).where(Brigade.id == brigade_id)
        )
        brigade = brigade_result.scalar_one_or_none()
        
        if not brigade:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Бригада с ID {brigade_id} не найдена"
            )
        
        # Получение объектов учета (для примера в dropdown)
        objects_result = await db.execute(
            select(CostObject).order_by(CostObject.name)
        )
        cost_objects = objects_result.scalars().all()
        
        # Создание Excel книги
        wb = Workbook()
        ws = wb.active
        ws.title = "Табель РТБ"
        
        # Стили
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=12)
        info_font = Font(size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Строка 1: Заголовок
        ws['A1'] = "Табель рабочего времени бригады"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Строка 2: Бригада
        ws['A2'] = "Бригада:"
        ws['A2'].font = info_font
        ws['B2'] = brigade.name
        ws['B2'].font = info_font
        
        # Строка 3: Период
        today = datetime.now().date()
        period_start = today.replace(day=1)
        period_end = (period_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        ws['A3'] = "Период:"
        ws['A3'].font = info_font
        ws['B3'] = f"{period_start.strftime('%d.%m.%Y')} - {period_end.strftime('%d.%m.%Y')}"
        ws['B3'].font = info_font
        
        # Строка 5: Заголовки таблицы
        headers = ["ФИО рабочего", "Дата (ДД.ММ.ГГГГ)", "Объект учета", "Часы"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Установка ширины колонок
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 10
        
        # Примеры заполнения (строки 6-8)
        example_data = [
            ["Иван Петров", "15.01.2026", cost_objects[0].name if cost_objects else "Объект 1", 8],
            ["Петр Иванов", "15.01.2026", cost_objects[1].name if len(cost_objects) > 1 else "Объект 2", 8],
            ["Сергей Сидоров", "16.01.2026", cost_objects[0].name if cost_objects else "Объект 1", 10],
        ]
        
        for row_idx, row_data in enumerate(example_data, 6):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='left' if col_idx != 4 else 'center', vertical='center')
                cell.border = border
                
                # Светлый фон для примеров
                if row_idx == 6:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                elif row_idx == 7:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Строки для заполнения (пустые)
        for row_idx in range(9, 35):
            for col_idx in range(1, 5):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
        
        # Примечание
        ws['A36'] = "Инструкция: Заполните таблицу выше данными. Дата должна быть в формате ДД.ММ.ГГГГ"
        ws['A36'].font = Font(italic=True, size=9, color="666666")
        ws.merge_cells('A36:D36')
        
        # Сохранение в BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Генерирование имени файла
        filename = f"Template_TimeSheet_Brigade_{brigade_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Ошибка при создании шаблона: {str(e)}"
        )


@router.post("/upload-excel", response_model=TimeSheetResponse, status_code=status.HTTP_201_CREATED)
async def upload_timesheet_excel(
    file: UploadFile = File(..., description="Excel файл табеля (.xlsx)"),
    brigade_id: int = Query(..., description="ID бригады"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.FOREMAN, UserRole.HR_MANAGER]))
):
    """
    Загрузка табеля из Excel файла
    
    - Доступно: FOREMAN, HR_MANAGER
    - Формат: .xlsx
    - Парсит данные и создает табель в статусе DRAFT
    
    Ожидаемая структура Excel:
    - Строка 2: Бригада: [название]
    - Строка 3: Период: ДД.ММ.ГГГГ - ДД.ММ.ГГГГ
    - Строки 6+: ФИО | Дата | Объект | Часы
    """
    # Проверка типа файла
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Поддерживаются только файлы .xlsx"
        )
    
    # Сохранение во временный файл
    temp_file_path = None
    try:
        # Создание временного файла
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            content = await file.read()
            temp_file.write(content)
            temp_file_path = temp_file.name
        
        # Парсинг Excel
        try:
            parser = TimeSheetExcelParser(temp_file_path)
            parsed_data = parser.parse()
        except TimeSheetExcelParseError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Ошибка парсинга Excel: {str(e)}"
            )
        
        # Проверка бригады
        service = TimeSheetService(db)
        brigade = await service.get_brigade_by_id(brigade_id)
        
        if not brigade:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Бригада {brigade_id} не найдена"
            )
        
        # Проверка прав (FOREMAN может только для своей бригады)
        if UserRole.FOREMAN in current_user.role and brigade.foreman_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Вы можете создавать табели только для своей бригады"
            )
        
        # Получить ID членов бригады по ФИО
        member_map = {}  # {full_name: member_id}
        for member in brigade.members:
            member_map[member.full_name] = member.id
        
        # Получить все объекты учета для поиска по названию
        objects_result = await db.execute(
            select(CostObject)
        )
        cost_objects = objects_result.scalars().all()
        
        # Создать маппинг: название → ID (case-insensitive)
        object_map = {}  # {normalized_name: cost_object_id}
        for obj in cost_objects:
            normalized_name = obj.name.strip().lower()
            object_map[normalized_name] = obj.id
        
        # Преобразование items для API
        items_for_api = []
        unknown_members = set()
        unknown_objects = set()
        
        for item in parsed_data['items']:
            member_name = item['member_name']
            
            # Поиск ID члена бригады
            member_id = member_map.get(member_name)
            if not member_id:
                unknown_members.add(member_name)
                continue
            
            # Поиск объекта по названию в БД
            cost_object_name = item['cost_object_name']
            cost_object_id = None
            
            if cost_object_name and cost_object_name != "Не указан":
                # Поиск с учетом регистра
                normalized_search = cost_object_name.strip().lower()
                cost_object_id = object_map.get(normalized_search)
                
                if not cost_object_id:
                    # Объект не найден - добавляем в unknown
                    unknown_objects.add(cost_object_name)
                    # Пропускаем эту запись (или используем fallback)
                    continue
            else:
                # Если объект не указан - пропускаем
                unknown_objects.add("(не указан)")
                continue
            
            items_for_api.append({
                'member_id': member_id,
                'date': item['date'].isoformat(),
                'cost_object_id': cost_object_id,
                'hours': float(item['hours'])
            })
        
        # Валидация
        if not items_for_api:
            error_msg = "Не найдено ни одного совпадения членов бригады"
            if unknown_members:
                error_msg += f". Неизвестные сотрудники: {', '.join(list(unknown_members)[:5])}"
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=error_msg
            )
        
        # Создание табеля через service
        timesheet_data = TimeSheetCreate(
            brigade_id=brigade_id,
            period_start=parsed_data['period_start'],
            period_end=parsed_data['period_end'],
            items=items_for_api
        )
        
        timesheet = await service.create_timesheet(timesheet_data, current_user.id)
        timesheet = await service.get_timesheet_by_id(timesheet.id)
        
        # Формирование ответа
        response = TimeSheetResponse(
            id=timesheet.id,
            brigade_id=timesheet.brigade_id,
            brigade_name=timesheet.brigade.name,
            period_start=timesheet.period_start,
            period_end=timesheet.period_end,
            status=timesheet.status.value,
            hour_rate=timesheet.hour_rate,
            total_hours=timesheet.total_hours,
            total_amount=timesheet.total_amount,
            items=[
                TimeSheetItemResponse(
                    id=item.id,
                    member_id=item.member_id,
                    member_name=item.member.full_name,
                    date=item.date,
                    cost_object_id=item.cost_object_id,
                    cost_object_name=item.cost_object.name,
                    hours=item.hours
                )
                for item in timesheet.items
            ],
            created_at=timesheet.created_at,
            updated_at=timesheet.updated_at
        )
        
        # Добавить предупреждения в ответ
        warnings = []
        if unknown_members:
            warnings.append(f"Пропущено сотрудников (не в бригаде): {len(unknown_members)}")
        if unknown_objects:
            warnings.append(f"Объекты не найдены в БД: {', '.join(list(unknown_objects)[:3])}")
        
        if warnings:
            # Логируем warnings для отладки
            print(f"[WARNING] Excel upload warnings: {warnings}")
        
        return response
    
    finally:
        # Удаление временного файла
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
            except:
                pass


@router.post("/upload-preview")
async def upload_excel_preview(
    file: UploadFile = File(..., description="Excel файл табеля (.xlsx)"),
    brigade_id: int = Query(..., description="ID бригады"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.FOREMAN.value, UserRole.HR_MANAGER.value]))
) -> Dict[str, Any]:
    """
    Предпросмотр табеля из Excel перед сохранением
    
    - Доступно: FOREMAN, HR_MANAGER
    - Формат: .xlsx
    - Возвращает preview_id и данные для проверки
    - Preview истекает через 1 час
    
    Ответ содержит:
    - preview_id — для confirm
    - items — все строки с валидацией
    - stats — статистика (валидные/невалидные, часы, переработки)
    - errors/warnings по каждой строке
    """
    # Проверка типа файла
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Поддерживаются только файлы .xlsx"
        )
    
    # Чтение файла
    content = await file.read()
    
    # Создание preview
    preview_service = ExcelPreviewService(db)
    
    try:
        preview_data = await preview_service.create_preview(
            file_content=content,
            brigade_id=brigade_id,
            user_id=current_user.id
        )
        
        return {
            "preview_id": preview_data['preview_id'],
            "brigade_id": preview_data['brigade_id'],
            "brigade_name": preview_data['brigade_name'],
            "period_start": preview_data['period_start'],
            "period_end": preview_data['period_end'],
            "stats": preview_data['stats'],
            "items": preview_data['items'],
            "expires_at": preview_data['expires_at'],
            "message": "Предпросмотр создан. Проверьте данные и подтвердите загрузку."
        }
    
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )


@router.post("/confirm-upload", response_model=TimeSheetResponse, status_code=status.HTTP_201_CREATED)
async def confirm_excel_upload(
    preview_id: str = Query(..., description="ID предпросмотра"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.FOREMAN.value, UserRole.HR_MANAGER.value]))
):
    """
    Подтверждение загрузки табеля после предпросмотра
    
    - Доступно: FOREMAN, HR_MANAGER
    - Использует данные из preview
    - Создаёт табель в статусе DRAFT
    - После создания удаляет preview
    """
    preview_service = ExcelPreviewService(db)
    
    # Получение preview
    preview_data = preview_service.get_preview(preview_id, current_user.id)
    
    if not preview_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Preview не найден или истёк. Загрузите файл заново."
        )
    
    # Проверка на наличие ошибок
    if preview_data['stats']['has_errors']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"В данных есть ошибки ({preview_data['stats']['invalid_rows']} строк). Исправьте файл и загрузите заново."
        )
    
    # Фильтрация только валидных строк
    valid_items = [
        {
            'member_id': item['member_id'],
            'date': item['date'],
            'cost_object_id': item['cost_object_id'],
            'hours': item['hours']
        }
        for item in preview_data['items']
        if item['valid']
    ]
    
    if not valid_items:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Нет валидных строк для создания табеля"
        )
    
    # Создание табеля
    service = TimeSheetService(db)
    
    try:
        timesheet_data = TimeSheetCreate(
            brigade_id=preview_data['brigade_id'],
            period_start=preview_data['period_start'],
            period_end=preview_data['period_end'],
            items=valid_items
        )
        
        timesheet = await service.create_timesheet(timesheet_data, current_user.id)
        timesheet = await service.get_timesheet_by_id(timesheet.id)
        
        # Удаление preview после успешного создания
        preview_service.delete_preview(preview_id)
        
        # Формирование ответа
        response = TimeSheetResponse(
            id=timesheet.id,
            brigade_id=timesheet.brigade_id,
            brigade_name=timesheet.brigade.name,
            period_start=timesheet.period_start,
            period_end=timesheet.period_end,
            status=timesheet.status,
            hour_rate=timesheet.hour_rate,
            total_hours=timesheet.total_hours,
            total_amount=timesheet.total_amount,
            items=[
                TimeSheetItemResponse(
                    id=item.id,
                    member_id=item.member_id,
                    member_name=item.member.full_name,
                    date=item.date,
                    cost_object_id=item.cost_object_id,
                    cost_object_name=item.cost_object.name,
                    hours=item.hours
                )
                for item in timesheet.items
            ],
            created_at=timesheet.created_at,
            updated_at=timesheet.updated_at
        )
        
        # Логирование в аудит
        from app.services.audit_service import AuditService
        audit_service = AuditService(db)
        await audit_service.log_action(
            user_id=current_user.id,
            action="CREATE_TIMESHEET_FROM_EXCEL",
            entity_type="TimeSheet",
            entity_id=timesheet.id,
            description=f"Создан табель #{timesheet.id} из Excel ({preview_data['stats']['valid_rows']} строк)"
        )
        
        return response
    
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )


@router.get("/preview/{preview_id}")
async def get_preview(
    preview_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Dict[str, Any]:
    """
    Получение существующего preview по ID
    
    - Доступно: всем авторизованным (только свои preview)
    - Возвращает данные preview если не истёк
    """
    preview_service = ExcelPreviewService(db)
    
    preview_data = preview_service.get_preview(preview_id, current_user.id)
    
    if not preview_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Preview не найден или истёк"
        )
    
    return preview_data

