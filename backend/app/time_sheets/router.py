"""Роутер для табелей рабочего времени (РТБ)"""
from datetime import date, datetime
from typing import List, Optional
import tempfile
import os
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO

from app.core.database import get_db
from app.auth.dependencies import require_roles, get_current_user
from app.core.models_base import UserRole, TimeSheetStatus
from app.models import User, Brigade
from app.time_sheets.service import TimeSheetService
from app.time_sheets.excel_parser import TimeSheetExcelParser, TimeSheetExcelParseError
from app.time_sheets.schemas import (
    TimeSheetCreate, TimeSheetResponse, TimeSheetListItem,
    TimeSheetSubmitRequest, TimeSheetApproveRequest,
    TimeSheetRejectRequest, TimeSheetItemResponse
)

router = APIRouter()


def get_status_key(status_val: Optional[str]) -> str:
    """Безопасное получение ключа статуса (DRAFT, SUBMITTED...)"""
    if not status_val:
        return "DRAFT"
    try:
        # Попытка найти по значению ("ЧЕРНОВИК" -> "DRAFT")
        return TimeSheetStatus(status_val).name
    except ValueError:
        # Попытка проверить, не является ли это уже ключом ("DRAFT")
        if status_val in TimeSheetStatus.__members__:
            return status_val
        return "DRAFT"

@router.post("/", response_model=TimeSheetResponse, status_code=status.HTTP_201_CREATED)
async def create_timesheet(
    data: TimeSheetCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.FOREMAN.value]))
):
    """
    Создание чернового табеля
    """
    service = TimeSheetService(db)
    
    # Проверка, что пользователь - бригадир и у него есть бригада
    from sqlalchemy import select
    brigade_query = select(Brigade).where(Brigade.foreman_id == current_user.id)
    brigade_result = await db.execute(brigade_query)
    brigade = brigade_result.scalar_one_or_none()
    
    if not brigade:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Пользователь не является бригадиром или не назначена бригада"
        )
    
    try:
        timesheet = await service.create_timesheet(
            brigade_id=brigade.id,
            period_start=data.period_start
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    
    return TimeSheetResponse(
        id=timesheet.id,
        brigade_id=timesheet.brigade_id,
        brigade_name=timesheet.brigade.name,
        period_start=timesheet.period_start,
        period_end=timesheet.period_end,
        status=get_status_key(timesheet.status),
        hour_rate=timesheet.hour_rate,
        total_hours=timesheet.total_hours,
        total_amount=timesheet.total_amount,
        items=[],
        created_at=timesheet.created_at,
        updated_at=timesheet.updated_at
    )


@router.get("/", response_model=List[TimeSheetListItem])
async def get_timesheets(
    status: Optional[str] = Query(None, description="Фильтр по статусу"),
    period_start: Optional[date] = Query(None, description="Начало периода"),
    period_end: Optional[date] = Query(None, description="Конец периода"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([
        UserRole.FOREMAN.value, 
        UserRole.HR_MANAGER.value, 
        UserRole.MANAGER.value, 
        UserRole.ADMIN.value
    ]))
):
    """
    Получение списка табелей
    
    - FOREMAN: только табели своей бригады
    - HR_MANAGER, MANAGER, ADMIN: все табели
    """
    service = TimeSheetService(db)
    
    # Преобразование статуса
    status_enum = None
    if status:
        try:
            status_enum = TimeSheetStatus(status)
        except ValueError:
            # Maybe it is a name?
            if status in TimeSheetStatus.__members__:
                 status_enum = TimeSheetStatus[status]
            else:
                 raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Некорректный статус: {status}"
                )
    
    # Определение прав доступа
    if UserRole.FOREMAN.value in current_user.roles:
        # Бригадир видит только свои табели
        from sqlalchemy import select
        brigade_query = select(Brigade).where(Brigade.foreman_id == current_user.id)
        brigade_result = await db.execute(brigade_query)
        brigade = brigade_result.scalar_one_or_none()
        
        if not brigade:
            return []
        
        timesheets = await service.get_timesheets_by_brigade(
            brigade.id, status_enum
        )
    else:
        # Менеджеры видят все табели
        timesheets = await service.get_all_timesheets(
            status_enum, period_start, period_end
        )
    
    
    result = []
    for ts in timesheets:
        # Foreman Name
        foreman_name = "Не назначен"
        if ts.brigade and ts.brigade.foreman:
            foreman_name = ts.brigade.foreman.full_name or ts.brigade.foreman.username
        
        # Objects Info
        objects = set()
        for item in ts.items:
            if item.cost_object:
                objects.add(item.cost_object.name)
        objects_str = ", ".join(sorted(list(objects)))
        
        result.append(
            TimeSheetListItem(
                id=ts.id,
                brigade_name=ts.brigade.name,
                foreman_name=foreman_name,
                objects_info=objects_str,
                period_start=ts.period_start,
                period_end=ts.period_end,
                status=get_status_key(ts.status),
                total_hours=ts.total_hours,
                created_at=ts.created_at
            )
        )
            
    return result


@router.get("/{timesheet_id}", response_model=TimeSheetResponse)
async def get_timesheet(
    timesheet_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Детальная информация о табеле
    
    - FOREMAN: только свои табели
    - HR_MANAGER, MANAGER, ADMIN: все табели
    """
    service = TimeSheetService(db)
    timesheet = await service.get_timesheet_by_id(timesheet_id)
    
    if not timesheet:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Табель {timesheet_id} не найден"
        )
    
    # Проверка прав доступа
    if UserRole.FOREMAN.value in current_user.roles:
        if timesheet.brigade.foreman_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Доступ запрещен"
            )
    
    return TimeSheetResponse(
        id=timesheet.id,
        brigade_id=timesheet.brigade_id,
        brigade_name=timesheet.brigade.name,
        period_start=timesheet.period_start,
        period_end=timesheet.period_end,
        status=get_status_key(timesheet.status),
        hour_rate=timesheet.hour_rate,
        total_hours=timesheet.total_hours,
        total_amount=timesheet.total_amount,
        items=[
            TimeSheetItemResponse(
                id=item.id,
                member_id=item.member_id,
                member_name=item.member.full_name,
                date=item.date,
                cost_object_id=item.cost_object_id,
                cost_object_name=item.cost_object.name,
                hours=item.hours
            )
            for item in timesheet.items
        ],
        created_at=timesheet.created_at,
        updated_at=timesheet.updated_at
    )


@router.post("/{timesheet_id}/submit", response_model=TimeSheetResponse)
async def submit_timesheet(
    timesheet_id: int,
    request: TimeSheetSubmitRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.FOREMAN]))
):
    """
    Отправка табеля на рассмотрение
    
    Переход: DRAFT -> UNDER_REVIEW
    """
    service = TimeSheetService(db)
    
    try:
        timesheet = await service.submit_timesheet(timesheet_id, current_user.id)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    
    # Получение детальной информации
    timesheet = await service.get_timesheet_by_id(timesheet.id)
    
    return TimeSheetResponse(
        id=timesheet.id,
        brigade_id=timesheet.brigade_id,
        brigade_name=timesheet.brigade.name,
        period_start=timesheet.period_start,
        period_end=timesheet.period_end,
        status=get_status_key(timesheet.status),
        hour_rate=timesheet.hour_rate,
        total_hours=timesheet.total_hours,
        total_amount=timesheet.total_amount,
        items=[
            TimeSheetItemResponse(
                id=item.id,
                member_id=item.member_id,
                member_name=item.member.full_name,
                date=item.date,
                cost_object_id=item.cost_object_id,
                cost_object_name=item.cost_object.name,
                hours=item.hours
            )
            for item in timesheet.items
        ],
        created_at=timesheet.created_at,
        updated_at=timesheet.updated_at
    )


@router.post("/{timesheet_id}/approve", response_model=TimeSheetResponse)
async def approve_timesheet(
    timesheet_id: int,
    request: TimeSheetApproveRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.HR_MANAGER]))
):
    """
    Утверждение табеля
    
    Переход: UNDER_REVIEW -> APPROVED
    Устанавливает ставку и создает записи затрат
    """
    service = TimeSheetService(db)
    
    try:
        timesheet = await service.approve_timesheet(timesheet_id, request.items)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    
    # Получение детальной информации
    timesheet = await service.get_timesheet_by_id(timesheet.id)
    
    return TimeSheetResponse(
        id=timesheet.id,
        brigade_id=timesheet.brigade_id,
        brigade_name=timesheet.brigade.name,
        period_start=timesheet.period_start,
        period_end=timesheet.period_end,
        status=get_status_key(timesheet.status),
        hour_rate=timesheet.hour_rate,
        total_hours=timesheet.total_hours,
        total_amount=timesheet.total_amount,
        items=[
            TimeSheetItemResponse(
                id=item.id,
                member_id=item.member_id,
                member_name=item.member.full_name,
                date=item.date,
                cost_object_id=item.cost_object_id,
                cost_object_name=item.cost_object.name,
                hours=item.hours
            )
            for item in timesheet.items
        ],
        created_at=timesheet.created_at,
        updated_at=timesheet.updated_at
    )


@router.post("/{timesheet_id}/reject", response_model=TimeSheetResponse)
async def reject_timesheet(
    timesheet_id: int,
    request: TimeSheetRejectRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.HR_MANAGER]))
):
    """
    Отклонение табеля
    
    Переход: UNDER_REVIEW -> CORRECTED
    """
    service = TimeSheetService(db)
    
    try:
        timesheet = await service.reject_timesheet(timesheet_id, request.comment)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    
    # Получение детальной информации
    timesheet = await service.get_timesheet_by_id(timesheet.id)
    
    return TimeSheetResponse(
        id=timesheet.id,
        brigade_id=timesheet.brigade_id,
        brigade_name=timesheet.brigade.name,
        period_start=timesheet.period_start,
        period_end=timesheet.period_end,
        status=get_status_key(timesheet.status),
        hour_rate=timesheet.hour_rate,
        total_hours=timesheet.total_hours,
        total_amount=timesheet.total_amount,
        items=[
            TimeSheetItemResponse(
                id=item.id,
                member_id=item.member_id,
                member_name=item.member.full_name,
                date=item.date,
                cost_object_id=item.cost_object_id,
                cost_object_name=item.cost_object.name,
                hours=item.hours
            )
            for item in timesheet.items
        ],
        created_at=timesheet.created_at,
        updated_at=timesheet.updated_at
    )

@router.get("/template")
async def download_excel_template():
    """
    Скачивание шаблона Excel для табеля РТБ
    
    - Доступно всем авторизованным пользователям
    - Возвращает готовый .xlsx файл с правильной структурой
    
    Структура шаблона:
    - Строка 1: Заголовок "Табель рабочего времени бригады"
    - Строка 2: Бригада: [название бригады]
    - Строка 3: Период: ДД.ММ.ГГГГ - ДД.ММ.ГГГГ
    - Строка 5: Заголовки столбцов (ФИО | Дата | Объект | Часы)
    - Строки 6+: Пустые строки для заполнения
    """
    # Создание workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Табель"
    
    # Стили
    header_font = Font(bold=True, size=14)
    label_font = Font(bold=True, size=11)
    column_header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Строка 1: Заголовок
    ws['A1'] = 'Табель рабочего времени бригады'
    ws['A1'].font = header_font
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Строка 2: Бригада
    ws['A2'] = 'Бригада:'
    ws['A2'].font = label_font
    ws['B2'] = '[Введите название бригады]'
    ws.merge_cells('B2:D2')
    
    # Строка 3: Период
    ws['A3'] = 'Период:'
    ws['A3'].font = label_font
    today = datetime.now().strftime('%d.%m.%Y')
    ws['B3'] = f'{today} - {today}'
    ws.merge_cells('B3:D3')
    
    # Строка 4: Пустая
    
    # Строка 5: Заголовки столбцов
    headers = ['ФИО', 'Дата', 'Объект', 'Часы']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = column_header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Добавление 10 пустых строк для заполнения
    for row in range(6, 16):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            # Добавление примеров в первую строку данных
            if row == 6:
                if col == 1:
                    cell.value = 'Иванов Иван Иванович'
                elif col == 2:
                    cell.value = today
                elif col == 3:
                    cell.value = 'Строительство дома №1'
                elif col == 4:
                    cell.value = 8
    
    # Установка ширины столбцов
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 10
    
    # Сохранение в BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Формирование имени файла
    filename = f'timesheet_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    return StreamingResponse(
        excel_buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@router.post("/{timesheet_id}/cancel")
async def cancel_timesheet(
    timesheet_id: int,
    cancellation_reason: str = Query(..., description="Причина отмены табеля"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.FOREMAN.value]))
):
    """
    Отмена табеля бригадиром
    
    - Доступно: FOREMAN (только для своих табелей)
    - Нельзя отменить утверждённый табель
    - Автоматически добавляется комментарий об отмене
    """
    service = TimeSheetService(db)
    
    try:
        timesheet = await service.cancel_timesheet(
            timesheet_id=timesheet_id,
            user_id=current_user.id,
            cancellation_reason=cancellation_reason
        )
        
        # Логирование в аудит
        from app.services.audit_service import AuditService
        audit_service = AuditService(db)
        await audit_service.log_action(
            user_id=current_user.id,
            action="CANCEL_TIMESHEET",
            entity_type="TimeSheet",
            entity_id=timesheet_id,
            old_value=None,
            new_value=TimeSheetStatus.CANCELLED.value,
            description=f"Отмена табеля #{timesheet_id}. Причина: {cancellation_reason}"
        )
        
        return {
            "id": timesheet.id,
            "status": timesheet.status,
            "cancellation_reason": timesheet.cancellation_reason,
            "message": f"Табель #{timesheet_id} отменён"
        }
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )


@router.post("/{timesheet_id}/comments")
async def add_comment(
    timesheet_id: int,
    comment_text: str = Query(..., description="Текст комментария"),
    comment_type: str = Query("GENERAL", description="Тип комментария"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.HR_MANAGER.value, UserRole.MANAGER.value]))
):
    """
    Добавление комментария к табелю
    
    - Доступно: HR_MANAGER, MANAGER
    - Типы комментариев: GENERAL, HR_CORRECTION, CANCELLATION
    - Автоматически уведомляет бригадира
    """
    service = TimeSheetService(db)
    
    try:
        comment = await service.add_comment(
            timesheet_id=timesheet_id,
            user_id=current_user.id,
            comment_text=comment_text,
            comment_type=comment_type
        )
        
        # Логирование в аудит
        from app.services.audit_service import AuditService
        audit_service = AuditService(db)
        await audit_service.log_action(
            user_id=current_user.id,
            action="ADD_COMMENT",
            entity_type="TimeSheet",
            entity_id=timesheet_id,
            description=f"Добавлен комментарий к табелю #{timesheet_id}"
        )
        
        return comment
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )


@router.get("/{timesheet_id}/comments")
async def get_comments(
    timesheet_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Получение всех комментариев к табелю
    
    - Доступно: всем авторизованным
    - Возвращает историю комментариев в хронологическом порядке
    """
    service = TimeSheetService(db)
    
    try:
        comments = await service.get_comments(timesheet_id)
        return {
            "timesheet_id": timesheet_id,
            "comments": comments,
            "total": len(comments)
        }
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(e)
        )


@router.get("/{timesheet_id}/overtime-check")
async def check_overtime(
    timesheet_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([
        UserRole.HR_MANAGER.value,
        UserRole.MANAGER.value,
        UserRole.FOREMAN.value
    ]))
):
    """
    Проверка переработок в табеле (>12 часов в день)
    
    - Доступно: HR_MANAGER, MANAGER, FOREMAN
    - Возвращает список случаев переработки
    - Используется для валидации перед утверждением
    """
    service = TimeSheetService(db)
    
    try:
        result = await service.validate_overtime(timesheet_id)
        return result
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(e)
        )


# Добавить endpoint для загрузки Excel
from app.time_sheets.excel_upload import router as excel_router
router.include_router(excel_router, tags=['time-sheets-excel'])

