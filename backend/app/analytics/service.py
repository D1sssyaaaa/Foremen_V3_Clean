"""Бизнес-логика модуля аналитики"""
from datetime import date
from decimal import Decimal
from typing import List, Optional
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
import io

from app.models import (
    CostEntry, CostObject, EquipmentOrder,
    MaterialRequest, MaterialCost
)


class AnalyticsService:
    """Сервис для аналитики и отчетности"""
    
    def __init__(self, db: AsyncSession):
        self.db = db
    
    async def get_object_costs(
        self,
        object_id: int,
        period_start: Optional[date] = None,
        period_end: Optional[date] = None
    ) -> dict:
        """
        Получение затрат по объекту
        
        Returns:
            Словарь с разбивкой затрат по типам
        """
        query = select(
            CostEntry.type,
            func.sum(CostEntry.amount).label('total')
        ).where(CostEntry.cost_object_id == object_id)
        
        if period_start:
            query = query.where(CostEntry.date >= period_start)
        
        if period_end:
            query = query.where(CostEntry.date <= period_end)
        
        query = query.group_by(CostEntry.type)
        
        result = await self.db.execute(query)
        costs = {row.type: row.total for row in result}
        
        return {
            'labor': costs.get('labor', Decimal('0')),
            'equipment': costs.get('equipment', Decimal('0')),
            'material': costs.get('material', Decimal('0')),
            'total': sum(costs.values())
        }
    
    async def get_all_objects_summary(
        self,
        period_start: Optional[date] = None,
        period_end: Optional[date] = None
    ) -> List[dict]:
        """Сводка по всем объектам"""
        # Получение всех объектов
        objects_query = select(CostObject)
        objects_result = await self.db.execute(objects_query)
        objects = objects_result.scalars().all()
        
        summary = []
        for obj in objects:
            costs = await self.get_object_costs(obj.id, period_start, period_end)
            
            # Расчет оставшегося бюджета
            remaining_budget = None
            budget_utilization = None
            
            if obj.contract_amount and obj.contract_amount > 0:
                remaining_budget = obj.contract_amount - costs['total']
                budget_utilization = (costs['total'] / obj.contract_amount) * 100
            
            summary.append({
                'object_id': obj.id,
                'object_name': obj.name,
                'total_labor_cost': costs['labor'],
                'total_equipment_cost': costs['equipment'],
                'total_material_cost': costs['material'],
                'total_cost': costs['total'],
                'contract_amount': obj.contract_amount,
                'remaining_budget': remaining_budget,
                'budget_utilization_percent': budget_utilization,
                'planned_labor_cost': obj.labor_amount,
                'planned_material_cost': obj.material_amount
            })
        
        return summary
    
    async def export_to_excel(
        self,
        period_start: date,
        period_end: date,
        object_id: Optional[int] = None
    ) -> bytes:
        """
        Экспорт аналитики затрат в Excel
        
        Args:
            period_start: Начало периода
            period_end: Конец периода
            object_id: ID объекта (если None - все объекты)
        
        Returns:
            Байты Excel файла
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl не установлен. Выполните: pip install openpyxl")
        
        # Получение данных
        summary = await self.get_all_objects_summary(period_start, period_end)
        
        if object_id:
            summary = [s for s in summary if s['object_id'] == object_id]
        
        # Создание Excel файла
        wb = Workbook()
        ws = wb.active
        ws.title = "Аналитика затрат"
        
        # Заголовок отчета
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"Аналитика затрат за период {period_start.strftime('%d.%m.%Y')} - {period_end.strftime('%d.%m.%Y')}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        
        ws.row_dimensions[1].height = 30
        
        # Заголовки столбцов
        headers = [
            'Объект', 'ФОТ', 'Материалы', 'Техника', 'Всего затрат', 
            'Договор', 'Остаток', 'Освоение %'
        ]
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[3].height = 25
        
        # Данные
        row_num = 4
        total_labor = Decimal('0')
        total_material = Decimal('0')
        total_equipment = Decimal('0')
        total_all = Decimal('0')
        total_contract = Decimal('0')
        total_remaining = Decimal('0')
        
        for item in summary:
            ws.cell(row=row_num, column=1, value=item['object_name'])
            ws.cell(row=row_num, column=2, value=float(item['total_labor_cost'] or 0))
            ws.cell(row=row_num, column=3, value=float(item['total_material_cost'] or 0))
            ws.cell(row=row_num, column=4, value=float(item['total_equipment_cost'] or 0))
            ws.cell(row=row_num, column=5, value=float(item['total_cost'] or 0))
            ws.cell(row=row_num, column=6, value=float(item['contract_amount'] or 0))
            ws.cell(row=row_num, column=7, value=float(item['remaining_budget'] or 0))
            
            if item['budget_utilization_percent']:
                ws.cell(row=row_num, column=8, value=f"{float(item['budget_utilization_percent']):.1f}%")
            
            # Накопление итогов (конвертируем float в Decimal)
            total_labor += Decimal(str(item['total_labor_cost'] or 0))
            total_material += Decimal(str(item['total_material_cost'] or 0))
            total_equipment += Decimal(str(item['total_equipment_cost'] or 0))
            total_all += Decimal(str(item['total_cost'] or 0))
            total_contract += Decimal(str(item['contract_amount'] or 0))
            total_remaining += Decimal(str(item['remaining_budget'] or 0))
            
            # Форматирование ячеек
            for col in range(1, 9):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                if col >= 2:  # Числовые колонки
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
            
            row_num += 1
        
        # Итоговая строка
        total_row = row_num
        ws.cell(row=total_row, column=1, value='ИТОГО')
        ws.cell(row=total_row, column=2, value=float(total_labor))
        ws.cell(row=total_row, column=3, value=float(total_material))
        ws.cell(row=total_row, column=4, value=float(total_equipment))
        ws.cell(row=total_row, column=5, value=float(total_all))
        ws.cell(row=total_row, column=6, value=float(total_contract))
        ws.cell(row=total_row, column=7, value=float(total_remaining))
        
        if total_contract > 0:
            ws.cell(row=total_row, column=8, value=f"{float((total_all / total_contract) * 100):.1f}%")
        
        # Форматирование итоговой строки
        total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        total_font = Font(bold=True)
        
        for col in range(1, 9):
            cell = ws.cell(row=total_row, column=col)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = border
            if col >= 2:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
        
        # Автоширина столбцов
        for col in range(1, 9):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = max(adjusted_width, 12)
        
        # Сохранение в байты
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    async def get_cost_breakdown(
        self,
        period_start: Optional[date] = None,
        period_end: Optional[date] = None
    ) -> List[dict]:
        """Разбивка затрат по типам"""
        query = select(
            CostEntry.type,
            func.sum(CostEntry.amount).label('total')
        )
        
        if period_start:
            query = query.where(CostEntry.date >= period_start)
        
        if period_end:
            query = query.where(CostEntry.date <= period_end)
        
        query = query.group_by(CostEntry.type)
        
        result = await self.db.execute(query)
        costs = {row.type: row.total for row in result}
        
        total = sum(costs.values())
        
        breakdown = []
        for cost_type, amount in costs.items():
            percentage = (amount / total * 100) if total > 0 else Decimal('0')
            breakdown.append({
                'type': cost_type,
                'amount': amount,
                'percentage': percentage
            })
        
        return breakdown
    
    async def get_object_detailed_report(
        self,
        object_id: int,
        period_start: Optional[date] = None,
        period_end: Optional[date] = None
    ) -> dict:
        """Детальный отчет по объекту"""
        # Получение объекта
        obj = await self.db.get(CostObject, object_id)
        if not obj:
            raise ValueError(f"Объект {object_id} не найден")
        
        # Затраты
        costs = await self.get_object_costs(object_id, period_start, period_end)
        
        # Часы труда (Legacy TimeSheet removed)
        labor_hours = Decimal('0')
        
        # Часы техники
        equipment_hours_query = select(func.sum(EquipmentOrder.total_hours)).where(
            EquipmentOrder.cost_object_id == object_id
        )
        equipment_hours_result = await self.db.execute(equipment_hours_query)
        equipment_hours = equipment_hours_result.scalar_one_or_none() or Decimal('0')
        
        # Количество документов
        timesheets_count = 0
        
        equipment_count_query = select(func.count(EquipmentOrder.id)).where(
            EquipmentOrder.cost_object_id == object_id
        )
        equipment_count = (await self.db.execute(equipment_count_query)).scalar_one()
        
        material_requests_query = select(func.count(MaterialRequest.id)).where(
            MaterialRequest.cost_object_id == object_id
        )
        material_requests_count = (await self.db.execute(material_requests_query)).scalar_one()
        
        upd_count_query = select(func.count(MaterialCost.id)).where(
            MaterialCost.cost_object_id == object_id
        )
        upd_count = (await self.db.execute(upd_count_query)).scalar_one()
        
        # Бюджет
        remaining_budget = None
        budget_utilization = None
        
        if obj.contract_amount and obj.contract_amount > 0:
            remaining_budget = obj.contract_amount - costs['total']
            budget_utilization = (costs['total'] / obj.contract_amount) * 100
        
        return {
            'object_id': obj.id,
            'object_name': obj.name,
            'contract_number': obj.contract_number,
            'contract_amount': obj.contract_amount,
            'start_date': obj.start_date,
            'end_date': obj.end_date,
            'total_labor_cost': costs['labor'],
            'total_labor_hours': labor_hours,
            'total_equipment_cost': costs['equipment'],
            'total_equipment_hours': equipment_hours,
            'total_material_cost': costs['material'],
            'total_cost': costs['total'],
            'remaining_budget': remaining_budget,
            'budget_utilization_percent': budget_utilization,
            'timesheets_count': timesheets_count,
            'equipment_orders_count': equipment_count,
            'material_requests_count': material_requests_count,
            'upd_documents_count': upd_count
        }
    
    async def get_cost_trend(
        self,
        object_id: int,
        period_start: date,
        period_end: date,
        granularity: str = 'day'  # day, week, month
    ) -> List[dict]:
        """Динамика затрат по объекту"""
        query = select(
            CostEntry.date,
            CostEntry.type,
            func.sum(CostEntry.amount).label('amount')
        ).where(
            CostEntry.cost_object_id == object_id,
            CostEntry.date >= period_start,
            CostEntry.date <= period_end
        ).group_by(CostEntry.date, CostEntry.type).order_by(CostEntry.date)
       
    async def get_top_5_objects(
        self,
        period_start: Optional[date] = None,
        period_end: Optional[date] = None,
        sort_by: str = "total_cost"
    ) -> List:
        """TOP-5 objects by cost or budget utilization"""
        from app.analytics.schemas import Top5Object
        from app.core.models_base import CostType
        
        # Get all objects with costs
        query = select(
            CostEntry.cost_object_id,
            func.sum(CostEntry.amount).label('total_cost')
        )
        
        if period_start:
            query = query.where(CostEntry.date >= period_start)
        if period_end:
            query = query.where(CostEntry.date <= period_end)
        
        query = query.group_by(CostEntry.cost_object_id)
        
        result = await self.db.execute(query)
        object_costs = {row.cost_object_id: row.total_cost for row in result.all()}
        
        if not object_costs:
            return []
        
        # Get objects
        result = await self.db.execute(
            select(CostObject).where(CostObject.id.in_(list(object_costs.keys())))
        )
        objects = result.scalars().all()
        
        # Calculate metrics
        top_objects = []
        for obj in objects:
            total_cost = object_costs[obj.id]
            contract_amount = obj.contract_amount or Decimal(0)
            
            budget_utilization = (
                (total_cost / contract_amount * 100)
                if contract_amount and contract_amount > 0
                else Decimal(0)
            )
            
            top_objects.append(
                Top5Object(
                    object_id=obj.id,
                    object_name=obj.name,
                    total_cost=total_cost,
                    contract_amount=contract_amount,
                    budget_utilization_percent=budget_utilization
                )
            )
        
        # Sort
        if sort_by == "budget_utilization":
            top_objects.sort(
                key=lambda x: x.budget_utilization_percent or Decimal(0),
                reverse=True
            )
        else:
            top_objects.sort(key=lambda x: x.total_cost, reverse=True)
        
        return top_objects[:5]
    
    async def get_cost_dynamics(
        self,
        object_id: Optional[int] = None,
        period_start: date = None,
        period_end: date = None,
        grouping: str = "month"
    ):
        """Cost dynamics by periods"""
        from app.analytics.schemas import CostDynamics, CostDynamicsPoint
        from app.core.models_base import CostType
        from collections import defaultdict
        from datetime import timedelta
        
        # Base query
        query = select(
            CostEntry.date,
            CostEntry.type,
            func.sum(CostEntry.amount).label('amount')
        )
        
        # Filters
        if object_id:
            query = query.where(CostEntry.cost_object_id == object_id)
        if period_start:
            query = query.where(CostEntry.date >= period_start)
        if period_end:
            query = query.where(CostEntry.date <= period_end)
        
        query = query.group_by(CostEntry.date, CostEntry.type)
        query = query.order_by(CostEntry.date)
        
        result = await self.db.execute(query)
        raw_data = result.all()
        
        # Aggregate by periods
        period_costs = defaultdict(lambda: {
            'labor': Decimal(0),
            'equipment': Decimal(0),
            'material': Decimal(0)
        })
        
        def get_period_key(dt: date) -> date:
            if grouping == "day":
                return dt
            elif grouping == "week":
                return dt - timedelta(days=dt.weekday())
            else:
                return dt.replace(day=1)
        
        # Grouping
        for row in raw_data:
            period_key = get_period_key(row.date)
            period_costs[period_key][row.type] += row.amount
        
        # Form points
        points = []
        for period_date in sorted(period_costs.keys()):
            costs = period_costs[period_date]
            total = costs['labor'] + costs['equipment'] + costs['material']
            
            points.append(
                CostDynamicsPoint(
                    period=period_date,
                    total_cost=total,
                    labor_cost=costs['labor'],
                    equipment_cost=costs['equipment'],
                    material_cost=costs['material']
                )
            )
        
        return CostDynamics(
            object_id=object_id,
            period_start=period_start,
            period_end=period_end,
            grouping=grouping,
            data_points=points
        )
    
    async def get_top_objects_by_deliveries(self, limit: int = 5) -> List[dict]:
        """
        Получение топ объектов по материалам (из УПД)
        
        Args:
            limit: Количество объектов (по умолчанию 5)
        
        Returns:
            Список объектов с суммами затрат на материалы
        """
        from app.models import MaterialCost
        
        # Используем MaterialCost (УПД) так как таблица Delivery пуста
        query = select(
            CostObject.id,
            CostObject.name,
            func.coalesce(func.sum(MaterialCost.total_amount), 0).label('deliveries')
        ).outerjoin(
            MaterialCost, CostObject.id == MaterialCost.cost_object_id
        ).group_by(
            CostObject.id, CostObject.name
        ).order_by(
            func.coalesce(func.sum(MaterialCost.total_amount), 0).desc()
        ).limit(limit)
        
        result = await self.db.execute(query)
        rows = result.all()
        
        return [
            {
                'name': row.name,
                'deliveries': float(row.deliveries)
            }
            for row in rows
        ]
    
    async def get_top_objects_by_equipment(self, limit: int = 5) -> List[dict]:
        """
        Получение топ объектов по расходам на спецтехнику
        
        Args:
            limit: Количество объектов (по умолчанию 5)
        
        Returns:
            Список объектов с суммами расходов на технику
        """
        from app.models import EquipmentCost
        
        query = select(
            CostObject.id,
            CostObject.name,
            func.coalesce(func.sum(EquipmentCost.total_amount), 0).label('equipment')
        ).outerjoin(
            EquipmentOrder, CostObject.id == EquipmentOrder.cost_object_id
        ).outerjoin(
            EquipmentCost, EquipmentOrder.id == EquipmentCost.equipment_order_id
        ).group_by(
            CostObject.id, CostObject.name
        ).order_by(
            func.coalesce(func.sum(EquipmentCost.total_amount), 0).desc()
        ).limit(limit)
        
        result = await self.db.execute(query)
        rows = result.all()
        
        return [
            {
                'name': row.name,
                'equipment': float(row.equipment)
            }
            for row in rows
        ]
