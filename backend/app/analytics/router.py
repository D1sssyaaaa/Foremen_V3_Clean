"""Р РѕСѓС‚РµСЂ РґР»СЏ Р°РЅР°Р»РёС‚РёРєРё Рё РѕС‚С‡РµС‚РЅРѕСЃС‚Рё"""
from datetime import date, datetime
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
import io
import csv
import json
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

from app.core.database import get_db
from app.auth.dependencies import require_roles, get_current_user
from app.core.models_base import UserRole
from app.models import User
from app.analytics.service import AnalyticsService
from app.analytics.schemas import (
    PeriodCostReport, ObjectDetailedReport,
    CostTrendReport, CostBreakdown, ObjectCostSummary,
    CostTrendItem, ExportRequest
)

router = APIRouter()


@router.get("/", response_model=list)
@router.get("/costs", response_model=list)
async def get_costs_analytics(
    period_start: date = Query(..., description="РќР°С‡Р°Р»Рѕ РїРµСЂРёРѕРґР°"),
    period_end: date = Query(..., description="РљРѕРЅРµС† РїРµСЂРёРѕРґР°"),
    object_id: Optional[int] = Query(None, description="Р¤РёР»СЊС‚СЂ РїРѕ РѕР±СЉРµРєС‚Сѓ"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.MANAGER, UserRole.ACCOUNTANT, UserRole.ADMIN]))
):
    """
    РђРЅР°Р»РёС‚РёРєР° Р·Р°С‚СЂР°С‚ Р·Р° РїРµСЂРёРѕРґ
    
    - Р Р°Р·Р±РёРІРєР° РїРѕ РѕР±СЉРµРєС‚Р°Рј Рё С‚РёРїР°Рј Р·Р°С‚СЂР°С‚
    - Р”РѕСЃС‚СѓРїРЅРѕ: MANAGER, ACCOUNTANT, ADMIN
    """
    service = AnalyticsService(db)
    
    # Р•СЃР»Рё СѓРєР°Р·Р°РЅ object_id - РІРµСЂРЅСѓС‚СЊ РґР°РЅРЅС‹Рµ РїРѕ РѕРґРЅРѕРјСѓ РѕР±СЉРµРєС‚Сѓ
    if object_id:
        from app.models import CostObject
        obj = await db.get(CostObject, object_id)
        if not obj:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"РћР±СЉРµРєС‚ {object_id} РЅРµ РЅР°Р№РґРµРЅ"
            )
        
        costs = await service.get_object_costs(object_id, period_start, period_end)
        
        return [{
            'object_id': obj.id,
            'object_name': obj.name,
            'period_start': period_start.isoformat(),
            'period_end': period_end.isoformat(),
            'labor_costs': float(costs.get('labor', 0)),
            'material_costs': float(costs.get('material', 0)),
            'equipment_costs': float(costs.get('equipment', 0)),
            'total_costs': float(costs.get('total', 0))
        }]
    
    # РРЅР°С‡Рµ - РїРѕ РІСЃРµРј РѕР±СЉРµРєС‚Р°Рј
    summary = await service.get_all_objects_summary(period_start, period_end)
    
    return [
        {
            'object_id': item['object_id'],
            'object_name': item['object_name'],
            'period_start': period_start.isoformat(),
            'period_end': period_end.isoformat(),
            'labor_costs': float(item['total_labor_cost'] or 0),
            'material_costs': float(item['total_material_cost'] or 0),
            'equipment_costs': float(item['total_equipment_cost'] or 0),
            'total_costs': float(item['total_cost'] or 0)
        }
        for item in summary
    ]


@router.get("/summary", response_model=PeriodCostReport)
async def get_cost_summary(
    period_start: Optional[date] = Query(None, description="РќР°С‡Р°Р»Рѕ РїРµСЂРёРѕРґР°"),
    period_end: Optional[date] = Query(None, description="РљРѕРЅРµС† РїРµСЂРёРѕРґР°"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.MANAGER, UserRole.ACCOUNTANT]))
):
    """
    РћР±С‰Р°СЏ СЃРІРѕРґРєР° Р·Р°С‚СЂР°С‚ Р·Р° РїРµСЂРёРѕРґ
    
    - Р Р°Р·Р±РёРІРєР° РїРѕ С‚РёРїР°Рј Р·Р°С‚СЂР°С‚
    - РЎРІРѕРґРєР° РїРѕ РІСЃРµРј РѕР±СЉРµРєС‚Р°Рј
    - Р”РѕСЃС‚СѓРїРЅРѕ: MANAGER, ACCOUNTANT
    """
    service = AnalyticsService(db)
    
    # РџРѕР»СѓС‡РµРЅРёРµ СЂР°Р·Р±РёРІРєРё Р·Р°С‚СЂР°С‚
    breakdown = await service.get_cost_breakdown(period_start, period_end)
    
    # РџРѕР»СѓС‡РµРЅРёРµ СЃРІРѕРґРєРё РїРѕ РѕР±СЉРµРєС‚Р°Рј
    objects_summary = await service.get_all_objects_summary(period_start, period_end)
    
    # Р Р°СЃС‡РµС‚ РѕР±С‰РµР№ СЃСѓРјРјС‹
    total_cost = sum(item['amount'] for item in breakdown)
    
    return PeriodCostReport(
        period_start=period_start or date(2020, 1, 1),
        period_end=period_end or date.today(),
        total_cost=total_cost,
        cost_breakdown=[
            CostBreakdown(**item) for item in breakdown
        ],
        objects_summary=[
            ObjectCostSummary(**obj) for obj in objects_summary
        ]
    )


@router.get("/objects/{object_id}", response_model=ObjectDetailedReport)
async def get_object_report(
    object_id: int,
    period_start: Optional[date] = Query(None, description="РќР°С‡Р°Р»Рѕ РїРµСЂРёРѕРґР°"),
    period_end: Optional[date] = Query(None, description="РљРѕРЅРµС† РїРµСЂРёРѕРґР°"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Р”РµС‚Р°Р»СЊРЅС‹Р№ РѕС‚С‡РµС‚ РїРѕ РѕР±СЉРµРєС‚Сѓ
    
    - Р—Р°С‚СЂР°С‚С‹ РїРѕ С‚РёРїР°Рј
    - Р§Р°СЃС‹ С‚СЂСѓРґР° Рё С‚РµС…РЅРёРєРё
    - РљРѕР»РёС‡РµСЃС‚РІРѕ РґРѕРєСѓРјРµРЅС‚РѕРІ
    - Р‘СЋРґР¶РµС‚ Рё РµРіРѕ РёСЃРїРѕР»СЊР·РѕРІР°РЅРёРµ
    """
    service = AnalyticsService(db)
    
    try:
        report = await service.get_object_detailed_report(
            object_id, period_start, period_end
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(e)
        )
    
    return ObjectDetailedReport(**report)


@router.get("/objects/{object_id}/trend", response_model=CostTrendReport)
async def get_cost_trend(
    object_id: int,
    period_start: date = Query(..., description="РќР°С‡Р°Р»Рѕ РїРµСЂРёРѕРґР°"),
    period_end: date = Query(..., description="РљРѕРЅРµС† РїРµСЂРёРѕРґР°"),
    granularity: str = Query("day", description="Р”РµС‚Р°Р»РёР·Р°С†РёСЏ: day, week, month"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Р”РёРЅР°РјРёРєР° Р·Р°С‚СЂР°С‚ РїРѕ РѕР±СЉРµРєС‚Сѓ
    
    - Р—Р°С‚СЂР°С‚С‹ РїРѕ РґРЅСЏРј/РЅРµРґРµР»СЏРј/РјРµСЃСЏС†Р°Рј
    - Р Р°Р·Р±РёРІРєР° РїРѕ С‚РёРїР°Рј
    """
    service = AnalyticsService(db)
    
    trend_data = await service.get_cost_trend(
        object_id, period_start, period_end, granularity
    )
    
    # РџРѕР»СѓС‡РµРЅРёРµ РёРјРµРЅРё РѕР±СЉРµРєС‚Р°
    from app.models import CostObject
    obj = await db.get(CostObject, object_id)
    if not obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"РћР±СЉРµРєС‚ {object_id} РЅРµ РЅР°Р№РґРµРЅ"
        )
    
    return CostTrendReport(
        object_id=object_id,
        object_name=obj.name,
        period_start=period_start,
        period_end=period_end,
        trend=[CostTrendItem(**item) for item in trend_data]
    )


@router.get("/export")
async def export_analytics(
    period_start: date = Query(..., description="РќР°С‡Р°Р»Рѕ РїРµСЂРёРѕРґР°"),
    period_end: date = Query(..., description="РљРѕРЅРµС† РїРµСЂРёРѕРґР°"),
    object_id: Optional[int] = Query(None, description="Р¤РёР»СЊС‚СЂ РїРѕ РѕР±СЉРµРєС‚Сѓ"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.MANAGER, UserRole.ACCOUNTANT, UserRole.ADMIN]))
):
    """
    Р­РєСЃРїРѕСЂС‚ Р°РЅР°Р»РёС‚РёРєРё РІ Excel
    
    - Р—Р°С‚СЂР°С‚С‹ РїРѕ РѕР±СЉРµРєС‚Р°Рј Р·Р° РїРµСЂРёРѕРґ
    - Р Р°Р·Р±РёРІРєР° РїРѕ С‚РёРїР°Рј Р·Р°С‚СЂР°С‚
    - Р”РѕСЃС‚СѓРїРЅРѕ: MANAGER, ACCOUNTANT, ADMIN
    """
    service = AnalyticsService(db)
    
    try:
        # РџРѕРїС‹С‚РєР° СЌРєСЃРїРѕСЂС‚РёСЂРѕРІР°С‚СЊ РІ Excel
        excel_file = await service.export_to_excel(period_start, period_end, object_id)
        
        return StreamingResponse(
            io.BytesIO(excel_file),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=analytics_{period_start}_{period_end}.xlsx'
            }
        )
    except Exception as e:
        logger.error(f"РћС€РёР±РєР° СЌРєСЃРїРѕСЂС‚Р° Excel: {str(e)}")
        
        # Fallback РЅР° CSV РµСЃР»Рё Excel РЅРµ СЂР°Р±РѕС‚Р°РµС‚
        summary = await service.get_all_objects_summary(period_start, period_end)
        
        if object_id:
            summary = [s for s in summary if s['object_id'] == object_id]
        
        output = io.StringIO()
        if summary:
            fieldnames = ['object_name', 'total_labor_cost', 'total_equipment_cost', 
                         'total_material_cost', 'total_cost', 'contract_amount', 
                         'remaining_budget', 'budget_utilization_percent']
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            
            for item in summary:
                row = {k: item.get(k, '') for k in fieldnames}
                writer.writerow(row)
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type='text/csv; charset=utf-8',
            headers={
                'Content-Disposition': f'attachment; filename=analytics_{period_start}_{period_end}.csv'
            }
        )


@router.post("/export")
async def export_data(
    request: ExportRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.MANAGER, UserRole.ACCOUNTANT]))
):
    """
    Р­РєСЃРїРѕСЂС‚ РґР°РЅРЅС‹С…
    
    - Р¤РѕСЂРјР°С‚С‹: csv, excel, json
    - Р¤РёР»СЊС‚СЂР°С†РёСЏ РїРѕ РѕР±СЉРµРєС‚Р°Рј Рё С‚РёРїР°Рј Р·Р°С‚СЂР°С‚
    """
    service = AnalyticsService(db)
    
    # РџРѕР»СѓС‡РµРЅРёРµ РґР°РЅРЅС‹С…
    if request.object_ids:
        # Р­РєСЃРїРѕСЂС‚ РїРѕ РєРѕРЅРєСЂРµС‚РЅС‹Рј РѕР±СЉРµРєС‚Р°Рј
        data = []
        for object_id in request.object_ids:
            try:
                report = await service.get_object_detailed_report(
                    object_id, request.period_start, request.period_end
                )
                data.append(report)
            except ValueError:
                continue
    else:
        # Р­РєСЃРїРѕСЂС‚ РІСЃРµС… РѕР±СЉРµРєС‚РѕРІ
        summary = await service.get_all_objects_summary(
            request.period_start, request.period_end
        )
        data = summary
    
    # Р¤РѕСЂРјРёСЂРѕРІР°РЅРёРµ РѕС‚РІРµС‚Р° РІ Р·Р°РІРёСЃРёРјРѕСЃС‚Рё РѕС‚ С„РѕСЂРјР°С‚Р°
    if request.format == 'csv':
        output = io.StringIO()
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            media_type='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename=report_{request.period_start}_{request.period_end}.csv'
            }
        )
    
    elif request.format == 'json':
        output = json.dumps(data, default=str, ensure_ascii=False, indent=2)
        
        return StreamingResponse(
            io.BytesIO(output.encode('utf-8')),
            media_type='application/json',
            headers={
                'Content-Disposition': f'attachment; filename=report_{request.period_start}_{request.period_end}.json'
            }
        )
    
    elif request.format == 'excel':
        # TODO: Р РµР°Р»РёР·РѕРІР°С‚СЊ СЌРєСЃРїРѕСЂС‚ РІ Excel (С‚СЂРµР±СѓРµС‚ openpyxl РёР»Рё xlsxwriter)
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Р­РєСЃРїРѕСЂС‚ РІ Excel РїРѕРєР° РЅРµ СЂРµР°Р»РёР·РѕРІР°РЅ"
        )
    
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"РќРµРїРѕРґРґРµСЂР¶РёРІР°РµРјС‹Р№ С„РѕСЂРјР°С‚: {request.format}"
        )


@router.get("/top-objects", response_model=list)
async def get_top_5_objects(
    period_start: Optional[date] = Query(None, description="РќР°С‡Р°Р»Рѕ РїРµСЂРёРѕРґР°"),
    period_end: Optional[date] = Query(None, description="РљРѕРЅРµС† РїРµСЂРёРѕРґР°"),
    sort_by: str = Query("total_cost", description="РЎРѕСЂС‚РёСЂРѕРІРєР°: total_cost РёР»Рё budget_utilization"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.MANAGER, UserRole.ACCOUNTANT]))
):
    """
    РўРћРџ-5 РѕР±СЉРµРєС‚РѕРІ РїРѕ Р·Р°С‚СЂР°С‚Р°Рј РёР»Рё РѕСЃРІРѕРµРЅРёСЋ Р±СЋРґР¶РµС‚Р°
    
    - sort_by: total_cost - РїРѕ РѕР±С‰РµР№ СЃСѓРјРјРµ Р·Р°С‚СЂР°С‚ (РїРѕ СѓРјРѕР»С‡Р°РЅРёСЋ)
    - sort_by: budget_utilization - РїРѕ РїСЂРѕС†РµРЅС‚Сѓ РѕСЃРІРѕРµРЅРёСЏ Р±СЋРґР¶РµС‚Р°
    - Р”РѕСЃС‚СѓРїРЅРѕ: MANAGER, ACCOUNTANT
    """
    from app.analytics.schemas import Top5Object
    service = AnalyticsService(db)
    
    # РџРѕР»СѓС‡РёС‚СЊ РўРћРџ-5
    top_objects = await service.get_top_5_objects(period_start, period_end, sort_by)
    
    return [Top5Object(**obj.__dict__) for obj in top_objects]


@router.get("/dynamics", response_model=dict)
async def get_cost_dynamics(
    object_id: Optional[int] = Query(None, description="ID РѕР±СЉРµРєС‚Р° (РµСЃР»Рё None - РїРѕ РІСЃРµРј РѕР±СЉРµРєС‚Р°Рј)"),
    period_start: date = Query(..., description="РќР°С‡Р°Р»Рѕ РїРµСЂРёРѕРґР°"),
    period_end: date = Query(..., description="РљРѕРЅРµС† РїРµСЂРёРѕРґР°"),
    grouping: str = Query("month", description="Р“СЂСѓРїРїРёСЂРѕРІРєР°: day, week, month"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.MANAGER, UserRole.ACCOUNTANT]))
):
    """
    Р”РёРЅР°РјРёРєР° Р·Р°С‚СЂР°С‚ РїРѕ РїРµСЂРёРѕРґР°Рј
    
    - grouping: day - РїРѕ РґРЅСЏРј, week - РїРѕ РЅРµРґРµР»СЏРј, month - РїРѕ РјРµСЃСЏС†Р°Рј
    - object_id: РµСЃР»Рё СѓРєР°Р·Р°РЅ - РїРѕ РєРѕРЅРєСЂРµС‚РЅРѕРјСѓ РѕР±СЉРµРєС‚Сѓ, РµСЃР»Рё None - РїРѕ РІСЃРµРј
    - Р”РѕСЃС‚СѓРїРЅРѕ: MANAGER, ACCOUNTANT
    """
    from app.analytics.schemas import CostDynamics
    service = AnalyticsService(db)
    
    dynamics = await service.get_cost_dynamics(
        object_id=object_id,
        period_start=period_start,
        period_end=period_end,
        grouping=grouping
    )
    
    return dynamics.__dict__


@router.get("/objects/{object_id}/costs", response_model=dict)
async def get_object_costs_detailed(
    object_id: int,
    period_start: Optional[date] = Query(None, description="РќР°С‡Р°Р»Рѕ РїРµСЂРёРѕРґР°"),
    period_end: Optional[date] = Query(None, description="РљРѕРЅРµС† РїРµСЂРёРѕРґР°"),
    cost_types: Optional[str] = Query(None, description="РўРёРїС‹ Р·Р°С‚СЂР°С‚ С‡РµСЂРµР· Р·Р°РїСЏС‚СѓСЋ (labor,equipment,material)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.MANAGER, UserRole.ACCOUNTANT, UserRole.FOREMAN]))
):
    """
    РђРіСЂРµРіРёСЂРѕРІР°РЅРЅС‹Рµ Р·Р°С‚СЂР°С‚С‹ РїРѕ РѕР±СЉРµРєС‚Сѓ СЃ С„РёР»СЊС‚СЂР°С†РёРµР№
    
    - Р Р°Р·Р±РёРІРєР° РїРѕ С‚РёРїР°Рј Р·Р°С‚СЂР°С‚
    - Р¤РёР»СЊС‚СЂ РїРѕ РїРµСЂРёРѕРґСѓ Рё С‚РёРїР°Рј
    - РРЅС„РѕСЂРјР°С†РёСЏ Рѕ Р±СЋРґР¶РµС‚Рµ
    - Р”РѕСЃС‚СѓРїРЅРѕ: MANAGER, ACCOUNTANT, FOREMAN
    """
    from app.analytics.schemas import ObjectCostSummary
    service = AnalyticsService(db)
    
    # РџР°СЂСЃРёРЅРі cost_types
    cost_types_list = None
    if cost_types:
        cost_types_list = [t.strip() for t in cost_types.split(',')]
    
    # РџРѕР»СѓС‡РёС‚СЊ РґР°РЅРЅС‹Рµ
    summary = await service.get_object_costs(
        object_id=object_id,
        period_start=period_start,
        period_end=period_end,
        cost_types=cost_types_list
    )
    
    if not summary:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"РћР±СЉРµРєС‚ {object_id} РЅРµ РЅР°Р№РґРµРЅ"
        )
    
    return summary.__dict__

@router.get("/export-excel", response_class=StreamingResponse)
async def export_analytics_to_excel(
    period_start: date = Query(..., description="РќР°С‡Р°Р»Рѕ РїРµСЂРёРѕРґР° (Р“Р“Р“Р“-РњРњ-Р”Р”)"),
    period_end: date = Query(..., description="РљРѕРЅРµС† РїРµСЂРёРѕРґР° (Р“Р“Р“Р“-РњРњ-Р”Р”)"),
    object_id: Optional[int] = Query(None, description="ID РѕР±СЉРµРєС‚Р° (РµСЃР»Рё None - РІСЃРµ РѕР±СЉРµРєС‚С‹)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.MANAGER, UserRole.ACCOUNTANT, UserRole.ADMIN]))
):
    """
    Р­РєСЃРїРѕСЂС‚ Р°РЅР°Р»РёС‚РёРєРё Р·Р°С‚СЂР°С‚ РІ Excel
    
    - Р›РёСЃС‚ 1: РЎРІРѕРґРєР° РїРѕ РѕР±СЉРµРєС‚Р°Рј
    - Р›РёСЃС‚ 2: Р”РёРЅР°РјРёРєР° Р·Р°С‚СЂР°С‚ РїРѕ РґР°С‚Р°Рј
    - Р›РёСЃС‚ 3: Р”РµС‚Р°Р»РёР·Р°С†РёСЏ РїРѕ РІРёРґР°Рј Р·Р°С‚СЂР°С‚
    - Р”РѕСЃС‚СѓРїРЅРѕ: MANAGER, ACCOUNTANT, ADMIN
    
    Р¤РѕСЂРјР°С‚ РІРѕР·РІСЂР°С‚Р°: .xlsx С„Р°Р№Р»
    """
    try:
        service = AnalyticsService(db)
        
        # РџРѕР»СѓС‡РµРЅРёРµ РґР°РЅРЅС‹С…
        if object_id:
            # Р•СЃР»Рё СѓРєР°Р·Р°РЅ object_id - СЌРєСЃРїРѕСЂС‚ РїРѕ РѕРґРЅРѕРјСѓ РѕР±СЉРµРєС‚Сѓ
            from app.models import CostObject
            obj = await db.get(CostObject, object_id)
            if not obj:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"РћР±СЉРµРєС‚ {object_id} РЅРµ РЅР°Р№РґРµРЅ"
                )
            
            costs = await service.get_object_costs(object_id, period_start, period_end)
            all_objects = [(obj, costs)]
        else:
            # РџРѕР»СѓС‡РёС‚СЊ РІСЃРµ РѕР±СЉРµРєС‚С‹ СЃ РёС… Р·Р°С‚СЂР°С‚Р°РјРё
            from app.models import CostObject
            from sqlalchemy import select
            result = await db.execute(select(CostObject).order_by(CostObject.name))
            objects = result.scalars().all()
            
            all_objects = []
            for obj in objects:
                costs = await service.get_object_costs(obj.id, period_start, period_end)
                all_objects.append((obj, costs))
        
        # РџРѕР»СѓС‡РµРЅРёРµ РґРёРЅР°РјРёРєРё
        try:
            dynamics = await service.get_cost_dynamics(
                object_id=object_id,
                period_start=period_start,
                period_end=period_end,
                grouping='day'
            )
        except:
            dynamics = None
        
        # РЎРѕР·РґР°РЅРёРµ Excel РєРЅРёРіРё
        wb = Workbook()
        
        # ===== Р›РРЎРў 1: РЎР’РћР”РљРђ РџРћ РћР‘РЄР•РљРўРђРњ =====
        ws_summary = wb.active
        ws_summary.title = "РЎРІРѕРґРєР° РїРѕ РѕР±СЉРµРєС‚Р°Рј"
        
        # РЎС‚РёР»Рё
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14)
        info_font = Font(size=10)
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Р—Р°РіРѕР»РѕРІРѕРє
        ws_summary['A1'] = "РђРЅР°Р»РёС‚РёРєР° Р·Р°С‚СЂР°С‚ РїРѕ РѕР±СЉРµРєС‚Р°Рј"
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:E1')
        ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # РџРµСЂРёРѕРґ
        ws_summary['A2'] = f"РџРµСЂРёРѕРґ: {period_start.strftime('%d.%m.%Y')} - {period_end.strftime('%d.%m.%Y')}"
        ws_summary['A2'].font = info_font
        ws_summary.merge_cells('A2:E2')
        
        # Р—Р°РіРѕР»РѕРІРєРё С‚Р°Р±Р»РёС†С‹
        headers = ["РћР±СЉРµРєС‚", "Р—Р°С‚СЂР°С‚С‹ С‚СЂСѓРґР°", "Р—Р°С‚СЂР°С‚С‹ РјР°С‚РµСЂРёР°Р»С‹", "Р—Р°С‚СЂР°С‚С‹ С‚РµС…РЅРёРєР°", "Р’РЎР•Р“Рћ"]
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Р”Р°РЅРЅС‹Рµ
        total_labor = 0
        total_material = 0
        total_equipment = 0
        total_all = 0
        
        for row_idx, (obj, costs) in enumerate(all_objects, 5):
            labor = float(costs.get('labor', 0))
            material = float(costs.get('material', 0))
            equipment = float(costs.get('equipment', 0))
            total = float(costs.get('total', 0))
            
            total_labor += labor
            total_material += material
            total_equipment += equipment
            total_all += total
            
            row_data = [obj.name, labor, material, equipment, total]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'right', vertical='center')
                
                if col_idx > 1:
                    cell.number_format = '#,##0.00 "в‚Ѕ"'
        
        # РС‚РѕРіРѕРІР°СЏ СЃС‚СЂРѕРєР°
        total_row = len(all_objects) + 5
        ws_summary.cell(row=total_row, column=1).value = "РРўРћР“Рћ:"
        ws_summary.cell(row=total_row, column=1).font = total_font
        
        total_row_data = [total_labor, total_material, total_equipment, total_all]
        for col_idx, value in enumerate(total_row_data, 2):
            cell = ws_summary.cell(row=total_row, column=col_idx)
            cell.value = value
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.number_format = '#,##0.00 "в‚Ѕ"'
        
        # РЁРёСЂРёРЅР° РєРѕР»РѕРЅРѕРє
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 18
        ws_summary.column_dimensions['C'].width = 18
        ws_summary.column_dimensions['D'].width = 18
        ws_summary.column_dimensions['E'].width = 18
        
        # ===== Р›РРЎРў 2: Р”РРќРђРњРРљРђ Р—РђРўР РђРў =====
        ws_dynamics = wb.create_sheet("Р”РёРЅР°РјРёРєР° РїРѕ РґР°С‚Р°Рј")
        
        ws_dynamics['A1'] = "Р”РёРЅР°РјРёРєР° Р·Р°С‚СЂР°С‚ РїРѕ РґР°С‚Р°Рј"
        ws_dynamics['A1'].font = title_font
        ws_dynamics.merge_cells('A1:D1')
        ws_dynamics['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        if dynamics and hasattr(dynamics, 'items') and dynamics.items:
            # Р—Р°РіРѕР»РѕРІРєРё
            dyn_headers = ["Р”Р°С‚Р°", "Р—Р°С‚СЂР°С‚С‹ С‚СЂСѓРґР°", "Р—Р°С‚СЂР°С‚С‹ РјР°С‚РµСЂРёР°Р»С‹", "Р’РЎР•Р“Рћ"]
            for col, header in enumerate(dyn_headers, 1):
                cell = ws_dynamics.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Р”Р°РЅРЅС‹Рµ
            for row_idx, item in enumerate(dynamics.items, 4):
                date_val = item.get('date', '')
                labor = float(item.get('labor_cost', 0))
                material = float(item.get('material_cost', 0))
                total = labor + material
                
                ws_dynamics.cell(row=row_idx, column=1).value = date_val
                ws_dynamics.cell(row=row_idx, column=2).value = labor
                ws_dynamics.cell(row=row_idx, column=3).value = material
                ws_dynamics.cell(row=row_idx, column=4).value = total
                
                for col in range(1, 5):
                    cell = ws_dynamics.cell(row=row_idx, column=col)
                    cell.border = border
                    if col > 1:
                        cell.number_format = '#,##0.00 "в‚Ѕ"'
                    cell.alignment = Alignment(horizontal='right' if col > 1 else 'center', vertical='center')
        else:
            ws_dynamics['A4'] = "Р”Р°РЅРЅС‹Рµ РЅРµ РЅР°Р№РґРµРЅС‹"
        
        # РЁРёСЂРёРЅР° РєРѕР»РѕРЅРѕРє
        ws_dynamics.column_dimensions['A'].width = 15
        ws_dynamics.column_dimensions['B'].width = 18
        ws_dynamics.column_dimensions['C'].width = 18
        ws_dynamics.column_dimensions['D'].width = 18
        
        # ===== Р›РРЎРў 3: Р”Р•РўРђР›РР—РђР¦РРЇ РџРћ Р’РР”РђРњ =====
        ws_details = wb.create_sheet("Р”РµС‚Р°Р»РёР·Р°С†РёСЏ РїРѕ РІРёРґР°Рј")
        
        ws_details['A1'] = "Р”РµС‚Р°Р»РёР·Р°С†РёСЏ Р·Р°С‚СЂР°С‚ РїРѕ РІРёРґР°Рј"
        ws_details['A1'].font = title_font
        ws_details.merge_cells('A1:C1')
        ws_details['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # РЎС‚Р°С‚РёСЃС‚РёРєР°
        stat_row = 3
        ws_details[f'A{stat_row}'] = "Р’РёРґ Р·Р°С‚СЂР°С‚"
        ws_details[f'B{stat_row}'] = "РЎСѓРјРјР° (в‚Ѕ)"
        ws_details[f'C{stat_row}'] = "% РѕС‚ РІСЃРµРіРѕ"
        
        for col in range(1, 4):
            cell = ws_details.cell(row=stat_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        stat_data = [
            ("Р—Р°С‚СЂР°С‚С‹ РЅР° С‚СЂСѓРґ", total_labor),
            ("Р—Р°С‚СЂР°С‚С‹ РЅР° РјР°С‚РµСЂРёР°Р»С‹", total_material),
            ("Р—Р°С‚СЂР°С‚С‹ РЅР° С‚РµС…РЅРёРєСѓ", total_equipment),
        ]
        
        for row_idx, (name, value) in enumerate(stat_data, stat_row + 1):
            ws_details.cell(row=row_idx, column=1).value = name
            ws_details.cell(row=row_idx, column=2).value = value
            
            if total_all > 0:
                percentage = (value / total_all) * 100
            else:
                percentage = 0
            
            ws_details.cell(row=row_idx, column=3).value = percentage
            
            for col in range(1, 4):
                cell = ws_details.cell(row=row_idx, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='left' if col == 1 else 'right', vertical='center')
                
                if col == 2:
                    cell.number_format = '#,##0.00 "в‚Ѕ"'
                elif col == 3:
                    cell.number_format = '0.00"%"'
        
        # РЁРёСЂРёРЅР° РєРѕР»РѕРЅРѕРє
        ws_details.column_dimensions['A'].width = 25
        ws_details.column_dimensions['B'].width = 18
        ws_details.column_dimensions['C'].width = 15
        
        # РЎРѕС…СЂР°РЅРµРЅРёРµ РІ BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Р“РµРЅРµСЂРёСЂРѕРІР°РЅРёРµ РёРјРµРЅРё С„Р°Р№Р»Р°
        filename = f"Analytics_Export_{period_start.strftime('%Y%m%d')}_to_{period_end.strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"РћС€РёР±РєР° РїСЂРё СЌРєСЃРїРѕСЂС‚Рµ Р°РЅР°Р»РёС‚РёРєРё: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"РћС€РёР±РєР° РїСЂРё СЃРѕР·РґР°РЅРёРё РѕС‚С‡РµС‚Р°: {str(e)}"
        )

@router.get("/top-objects-by-deliveries")
async def get_top_objects_by_deliveries(
    limit: int = Query(5, ge=1, le=20, description="оличество объектов"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.MANAGER, UserRole.ACCOUNTANT, UserRole.ADMIN]))
):
    """
    Топ объектов по доставкам материалов
    
    - озвращает топ N объектов с наибольшими суммами доставок
    - оступно: MANAGER, ACCOUNTANT, ADMIN
    """
    service = AnalyticsService(db)
    return await service.get_top_objects_by_deliveries(limit=limit)


@router.get("/top-objects-by-equipment")
async def get_top_objects_by_equipment(
    limit: int = Query(5, ge=1, le=20, description="оличество объектов"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.MANAGER, UserRole.ACCOUNTANT, UserRole.ADMIN]))
):
    """
    Топ объектов по расходам на спецтехнику
    
    - озвращает топ N объектов с наибольшими расходами на технику
    - оступно: MANAGER, ACCOUNTANT, ADMIN
    """
    service = AnalyticsService(db)
    return await service.get_top_objects_by_equipment(limit=limit)


