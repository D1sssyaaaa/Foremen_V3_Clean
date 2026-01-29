"""Создание тестового Excel файла табеля"""
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def create_test_timesheet_excel(filepath: str = "test_timesheet.xlsx"):
    """Создать тестовый Excel файл табеля"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Табель"
    
    # Заголовок
    ws['A1'] = "ТАБЕЛЬ УЧЕТА РАБОЧЕГО ВРЕМЕНИ БРИГАДЫ"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Бригада
    ws['A2'] = "Бригада: Отделочники №5"
    ws['A2'].font = Font(bold=True)
    
    # Период
    start = date.today() - timedelta(days=7)
    end = date.today()
    ws['A3'] = f"Период: {start.strftime('%d.%m.%Y')} - {end.strftime('%d.%m.%Y')}"
    ws['A3'].font = Font(bold=True)
    
    # Пустая строка
    ws['A4'] = ""
    
    # Заголовки таблицы
    headers = ['ФИО', 'Дата', 'Объект', 'Часы']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Данные (3 работника, 3 дня каждый)
    members = [
        "Иванов Иван Иванович",
        "Петров Петр Петрович",
        "Сидоров Сидор Сидорович"
    ]
    
    objects = [
        "ЖК Солнечный",
        "ЖК Центральный",
        "ЖК Морской"
    ]
    
    row = 6
    for i in range(3):  # 3 дня
        work_date = start + timedelta(days=i)
        for member_idx, member in enumerate(members):
            ws.cell(row=row, column=1, value=member)
            ws.cell(row=row, column=2, value=work_date.strftime('%d.%m.%Y'))
            ws.cell(row=row, column=3, value=objects[member_idx % 3])
            ws.cell(row=row, column=4, value=8)  # 8 часов
            row += 1
    
    # Автоширина колонок
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    
    # Сохранить
    wb.save(filepath)
    print(f"✓ Создан тестовый файл: {filepath}")
    print(f"  - Бригада: Отделочники №5")
    print(f"  - Период: {start.strftime('%d.%m.%Y')} - {end.strftime('%d.%m.%Y')}")
    print(f"  - Сотрудников: {len(members)}")
    print(f"  - Записей: {len(members) * 3}")


if __name__ == "__main__":
    create_test_timesheet_excel()
