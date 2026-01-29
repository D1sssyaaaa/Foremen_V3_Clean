"""Интеграционные тесты для API табелей"""
import pytest
from datetime import date, timedelta
from decimal import Decimal


def test_create_timesheet(client, auth_headers, test_foreman, test_brigade, test_object):
    """Тест создания табеля через API"""
    # Подготовить данные
    start_date = date.today()
    end_date = start_date + timedelta(days=2)
    
    members = test_brigade.members
    
    payload = {
        "brigade_id": test_brigade.id,
        "period_start": start_date.isoformat(),
        "period_end": end_date.isoformat(),
        "items": [
            {
                "member_id": members[0].id,
                "date": start_date.isoformat(),
                "cost_object_id": test_object.id,
                "hours": 8
            },
            {
                "member_id": members[1].id,
                "date": start_date.isoformat(),
                "cost_object_id": test_object.id,
                "hours": 8
            }
        ]
    }
    
    # Выполнить запрос
    response = client.post("/api/v1/time-sheets/", json=payload, headers=auth_headers)
    
    # Проверить
    assert response.status_code == 201
    data = response.json()
    
    assert data['brigade_id'] == test_brigade.id
    assert data['status'] == 'DRAFT'
    assert len(data['items']) == 2
    assert data['total_hours'] == 16


def test_submit_timesheet(client, auth_headers, test_foreman, test_brigade, test_object):
    """Тест отправки табеля на проверку"""
    # Создать табель
    start_date = date.today()
    members = test_brigade.members
    
    create_payload = {
        "brigade_id": test_brigade.id,
        "period_start": start_date.isoformat(),
        "period_end": (start_date + timedelta(days=1)).isoformat(),
        "items": [
            {
                "member_id": members[0].id,
                "date": start_date.isoformat(),
                "cost_object_id": test_object.id,
                "hours": 8
            }
        ]
    }
    
    create_response = client.post("/api/v1/time-sheets/", json=create_payload, headers=auth_headers)
    timesheet_id = create_response.json()['id']
    
    # Отправить на проверку
    submit_response = client.post(
        f"/api/v1/time-sheets/{timesheet_id}/submit",
        headers=auth_headers
    )
    
    assert submit_response.status_code == 200
    assert submit_response.json()['status'] == 'SUBMITTED'


def test_approve_timesheet(client, auth_headers, test_foreman, test_brigade, test_object, db_session):
    """Тест утверждения табеля"""
    # Создать HR менеджера
    from app.models import User
    from app.core.models_base import UserRole
    
    hr_manager = User(
        username="hr_manager",
        phone="+79009999999",
        role=[UserRole.HR_MANAGER],
        hashed_password="$2b$12$fake"
    )
    db_session.add(hr_manager)
    db_session.commit()
    
    hr_headers = {"Authorization": f"Bearer test_token_{hr_manager.id}"}
    
    # Создать и отправить табель
    start_date = date.today()
    members = test_brigade.members
    
    create_payload = {
        "brigade_id": test_brigade.id,
        "period_start": start_date.isoformat(),
        "period_end": (start_date + timedelta(days=1)).isoformat(),
        "items": [
            {
                "member_id": members[0].id,
                "date": start_date.isoformat(),
                "cost_object_id": test_object.id,
                "hours": 8
            }
        ]
    }
    
    create_response = client.post("/api/v1/time-sheets/", json=create_payload, headers=auth_headers)
    timesheet_id = create_response.json()['id']
    
    client.post(f"/api/v1/time-sheets/{timesheet_id}/submit", headers=auth_headers)
    
    # Утвердить
    approve_payload = {
        "hour_rate": 500.00
    }
    
    approve_response = client.post(
        f"/api/v1/time-sheets/{timesheet_id}/approve",
        json=approve_payload,
        headers=hr_headers
    )
    
    assert approve_response.status_code == 200
    data = approve_response.json()
    assert data['status'] == 'APPROVED'
    assert data['hour_rate'] == 500.00
    assert data['total_amount'] == 4000.00  # 8 часов * 500


def test_get_timesheets_list(client, auth_headers, test_foreman, test_brigade, test_object):
    """Тест получения списка табелей с фильтрацией"""
    # Создать несколько табелей
    start_date = date.today()
    members = test_brigade.members
    
    for i in range(3):
        payload = {
            "brigade_id": test_brigade.id,
            "period_start": (start_date + timedelta(days=i*7)).isoformat(),
            "period_end": (start_date + timedelta(days=i*7+1)).isoformat(),
            "items": [
                {
                    "member_id": members[0].id,
                    "date": (start_date + timedelta(days=i*7)).isoformat(),
                    "cost_object_id": test_object.id,
                    "hours": 8
                }
            ]
        }
        client.post("/api/v1/time-sheets/", json=payload, headers=auth_headers)
    
    # Получить список
    response = client.get("/api/v1/time-sheets/", headers=auth_headers)
    
    assert response.status_code == 200
    data = response.json()
    assert len(data) >= 3


def test_excel_parser_integration(test_brigade, test_object):
    """Интеграционный тест парсера Excel"""
    from app.time_sheets.excel_parser import TimeSheetExcelParser
    from openpyxl import Workbook
    import tempfile
    import os
    
    # Создать тестовый Excel
    wb = Workbook()
    ws = wb.active
    
    ws['A1'] = "ТАБЕЛЬ УЧЕТА РАБОЧЕГО ВРЕМЕНИ БРИГАДЫ"
    ws['A2'] = f"Бригада: {test_brigade.name}"
    ws['A3'] = f"Период: {date.today().strftime('%d.%m.%Y')} - {(date.today() + timedelta(days=2)).strftime('%d.%m.%Y')}"
    ws['A5'] = "ФИО"
    ws['B5'] = "Дата"
    ws['C5'] = "Объект"
    ws['D5'] = "Часы"
    
    members = test_brigade.members
    ws['A6'] = members[0].full_name
    ws['B6'] = date.today().strftime('%d.%m.%Y')
    ws['C6'] = test_object.name
    ws['D6'] = 8
    
    # Сохранить во временный файл
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        temp_path = tmp.name
    
    try:
        # Парсить
        parser = TimeSheetExcelParser(temp_path)
        data = parser.parse()
        
        assert data['brigade_name'] == test_brigade.name
        assert len(data['items']) == 1
        assert data['items'][0]['member_name'] == members[0].full_name
        assert data['items'][0]['hours'] == 8
    finally:
        os.unlink(temp_path)


def test_timesheet_validation(client, auth_headers, test_foreman, test_brigade, test_object):
    """Тест валидации данных табеля"""
    start_date = date.today()
    
    # Попытка создать табель с невалидными данными (более 24 часов)
    payload = {
        "brigade_id": test_brigade.id,
        "period_start": start_date.isoformat(),
        "period_end": (start_date + timedelta(days=1)).isoformat(),
        "items": [
            {
                "member_id": test_brigade.members[0].id,
                "date": start_date.isoformat(),
                "cost_object_id": test_object.id,
                "hours": 25  # Невалидно
            }
        ]
    }
    
    response = client.post("/api/v1/time-sheets/", json=payload, headers=auth_headers)
    
    # Должна быть ошибка валидации
    assert response.status_code == 422 or response.status_code == 400
